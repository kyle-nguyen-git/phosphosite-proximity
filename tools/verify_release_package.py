"""Verify technical release artifacts without standing in for human sign-off."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import tarfile
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from release_policy import is_source_workbook, sha256, source_tree_fingerprint


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "phase0_5" / "results"
REPORT = ROOT / "release" / "release_readiness_report.json"
EXPECTED_SHEETS = [
    "Summary", "Arm Estimates", "Arm Cutoffs", "Arm Descriptives",
    "Confidence Strata", "Regression Models", "Logistic Diagnostics",
    "Continuous Outcomes", "Predictor Benchmark", "Feature Definitions",
    "Feature Evidence", "Primary Analysis Data", "Inclusive Analysis Data",
    "Replicate Audit", "PAE Manifest", "Mismatch Audit", "Cohort Sensitivity",
    "PAE at 10A", "PAE Filter Grid", "SIFT Comparator", "Within Protein",
    "Residue Classes",
]
CSV_SHEETS = {
    "Arm Estimates": "cohort_arm_primary_estimates.csv",
    "Arm Cutoffs": "cohort_arm_cutoffs.csv",
    "Arm Descriptives": "cohort_arm_descriptives.csv",
    "Confidence Strata": "confidence_strata.csv",
    "Regression Models": "regression_models.csv",
    "Logistic Diagnostics": "logistic_fit_diagnostics.csv",
    "Continuous Outcomes": "continuous_outcomes.csv",
    "Predictor Benchmark": "predictor_benchmark.csv",
    "Feature Definitions": "feature_definition_sensitivity.csv",
    "Feature Evidence": "feature_evidence_audit.csv",
    "Primary Analysis Data": "phase0_5_primary_analysis.csv",
    "Inclusive Analysis Data": "phase0_5_inclusive_sensitivity_analysis.csv",
    "Replicate Audit": "replicate_aggregation_audit.csv",
    "PAE Manifest": "alphafold_pae_manifest.csv",
    "Mismatch Audit": "residue_mismatch_audit.csv",
    "Cohort Sensitivity": "cohort_sensitivity.csv",
    "PAE at 10A": "pae_column_sensitivity_at_10A.csv",
    "PAE Filter Grid": "pae_filter_grid_72x3.csv",
    "SIFT Comparator": "sift_comparator_sensitivity.csv",
    "Within Protein": "within_protein_auc.csv",
    "Residue Classes": "residue_class_sensitivity.csv",
}
FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


def frozen_hashes() -> dict[str, str]:
    return dict(re.findall(
        r"^- `([^`]+)` — `([0-9a-f]{64})`$",
        (ROOT / "NUMBERS.md").read_text(),
        re.MULTILINE,
    ))


def cff_field(text: str, name: str) -> str | None:
    match = re.search(rf"^{re.escape(name)}:\s*[\"']?([^\"'\n]+)", text, re.MULTILINE)
    return match.group(1).strip() if match else None


def verify_workbook() -> dict[str, object]:
    path = RESULTS / "phase0_5_supplement.xlsx"
    workbook = load_workbook(path, data_only=False, read_only=False)
    formula_cells: list[str] = []
    error_cells: list[str] = []
    row_checks: dict[str, bool] = {}
    table_count = 0
    for sheet in workbook.worksheets:
        table_count += len(sheet.tables)
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    formula_cells.append(f"{sheet.title}!{cell.coordinate}")
                if isinstance(cell.value, str) and any(
                    token in cell.value for token in FORMULA_ERRORS
                ):
                    error_cells.append(f"{sheet.title}!{cell.coordinate}")
        if sheet.title in CSV_SHEETS:
            with (RESULTS / CSV_SHEETS[sheet.title]).open(newline="") as handle:
                expected_rows = sum(1 for _ in csv.reader(handle))
            row_checks[sheet.title] = sheet.max_row == expected_rows
    return {
        "path": str(path.relative_to(ROOT)),
        "sha256": sha256(path),
        "sheet_names_exact": workbook.sheetnames == EXPECTED_SHEETS,
        "table_count": table_count,
        "one_table_per_sheet": table_count == len(EXPECTED_SHEETS),
        "csv_row_counts_match": all(row_checks.values()),
        "csv_row_checks": row_checks,
        "formula_cells": formula_cells,
        "error_cells": error_cells,
    }


def verify_archive(path: Path) -> dict[str, object]:
    forbidden: list[str] = []
    unsafe: list[str] = []
    symlinks: list[str] = []
    payloads: dict[str, bytes] = {}
    with tarfile.open(path, "r:gz") as archive:
        members = archive.getmembers()
        roots = {Path(member.name).parts[0] for member in members if member.name}
        for member in members:
            parts = Path(member.name).parts
            if member.name.startswith("/") or ".." in parts:
                unsafe.append(member.name)
            if member.issym() or member.islnk():
                symlinks.append(member.name)
            if not member.isfile():
                continue
            relative = Path(*parts[1:]) if len(parts) > 1 else Path(member.name)
            if (
                is_source_workbook(relative)
                or relative.name == "PMC7612524_supplementary.zip"
                or relative.as_posix() == "data/EMS132528-supplement-Supplementary_Information.pdf"
            ):
                forbidden.append(member.name)
            extracted = archive.extractfile(member)
            if extracted is not None:
                payloads[member.name] = extracted.read()
    checksum_names = [name for name in payloads if name.endswith("/SHA256SUMS")]
    checksum_failures: list[str] = []
    if len(checksum_names) == 1:
        prefix = checksum_names[0].rsplit("/", 1)[0]
        for line in payloads[checksum_names[0]].decode("utf-8").splitlines():
            expected, relative = line.split("  ", 1)
            member_name = f"{prefix}/{relative}"
            observed = hashlib.sha256(payloads.get(member_name, b"")).hexdigest()
            if observed != expected:
                checksum_failures.append(relative)
    else:
        checksum_failures.append("SHA256SUMS missing or duplicated")
    metadata_names = [name for name in payloads if name.endswith("/PACKAGE_METADATA.json")]
    package_metadata = (
        json.loads(payloads[metadata_names[0]].decode("utf-8"))
        if len(metadata_names) == 1 else {}
    )
    package_manifest_names = [name for name in payloads if name.endswith("/PACKAGE_MANIFEST.csv")]
    live_manifest_failures: list[str] = []
    if len(package_manifest_names) == 1:
        rows = list(csv.DictReader(io.StringIO(
            payloads[package_manifest_names[0]].decode("utf-8")
        )))
        for row in rows:
            relative = row["path"]
            if relative == "PACKAGE_METADATA.json":
                continue
            live_path = ROOT / relative
            if not live_path.is_file() or sha256(live_path) != row["sha256"]:
                live_manifest_failures.append(relative)
    else:
        live_manifest_failures.append("PACKAGE_MANIFEST.csv missing or duplicated")
    live_fingerprint = source_tree_fingerprint(ROOT)
    return {
        "path": str(path),
        "sha256": sha256(path),
        "one_top_level_directory": len(roots) == 1,
        "unsafe_members": unsafe,
        "symlink_members": symlinks,
        "forbidden_source_members": forbidden,
        "checksum_failures": checksum_failures,
        "package_metadata": package_metadata,
        "live_source_tree_fingerprint": live_fingerprint,
        "source_tree_fingerprint_matches_live_root": (
            package_metadata.get("source_tree_fingerprint_excluding_generated_outputs")
            == live_fingerprint
        ),
        "live_manifest_failures": live_manifest_failures,
        "file_count": len(payloads),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--archive", type=Path)
    args = parser.parse_args()

    frozen = frozen_hashes()
    required_frozen = {
        "results/statistics.json",
        "results/analysis_final.csv",
        "phase0_5/results/phase0_5_statistics.json",
    }
    freeze_checks = {
        relative: relative in frozen and sha256(ROOT / relative) == frozen[relative]
        for relative in required_frozen
    }
    scientific = json.loads((RESULTS / "verification_report.json").read_text())
    cache = json.loads((ROOT / "release_metadata" / "cache_verification_report.json").read_text())
    cache_manifest_rows = list(csv.DictReader(
        (ROOT / "release_metadata" / "third_party_data_manifest.csv").open()
    ))
    workbook = verify_workbook()
    pdf = ROOT / "manuscript" / "preprint_draft_v1.pdf"
    render = json.loads(
        (ROOT / "manuscript" / "rendered" / "render_manifest.json").read_text()
    )
    root_cff = (ROOT / "CITATION.cff").read_text()
    phase_cff = (ROOT / "phase0_5" / "CITATION.cff").read_text()
    review_files = [
        "statistical_methods_review.md",
        "biological_structural_review.md",
        "manuscript_claims_review.md",
        "reproducibility_release_review.md",
        "RESPONSE_LOG.md",
    ]
    required_docs = [
        "README.md", "NUMBERS.md", "SOURCE_RETRIEVAL.md", "THIRD_PARTY_NOTICES.md",
        "LICENSES.md", "sources.lock.json", "requirements-lock.txt", "reproduce.sh",
        "release/AUTHOR_SIGNOFF.md", "release/EXTERNAL_METHODS_REVIEW.md",
        "release/RELEASE_CHECKLIST.md", "release/DEPOSITION_CHECKLIST.md",
    ]
    checks = {
        "frozen_hashes_match_numbers": all(freeze_checks.values()),
        "scientific_verifier_passes": (
            scientific.get("status") == "PASS"
            and not scientific.get("failed")
            and all(scientific.get("checks", {}).values())
        ),
        "third_party_cache_verifier_passes": (
            cache.get("status") == "PASS"
            and cache.get("manifest_file_count") == len(cache_manifest_rows)
            and cache.get("all_file_hashes_match") is True
            and cache.get("all_sequences_versions_and_pae_dimensions_match") is True
        ),
        "workbook_sheet_names_exact": bool(workbook["sheet_names_exact"]),
        "workbook_one_table_per_sheet": bool(workbook["one_table_per_sheet"]),
        "workbook_rows_match_csvs": bool(workbook["csv_row_counts_match"]),
        "workbook_has_no_formulas_or_error_tokens": (
            not workbook["formula_cells"] and not workbook["error_cells"]
        ),
        "pdf_and_render_manifest_agree": (
            render.get("pdf_sha256") == sha256(pdf)
            and render.get("page_count") == len(PdfReader(str(pdf)).pages)
        ),
        "requirements_indirection_is_consistent": (
            (ROOT / "requirements.txt").read_text().strip() == "-r requirements-lock.txt"
            and (ROOT / "phase0_5" / "requirements-lock.txt").read_text().strip()
            == "-r ../requirements-lock.txt"
        ),
        "required_release_documents_exist": all((ROOT / path).is_file() for path in required_docs),
        "citation_metadata_shared_fields_agree_and_are_provisional": (
            all(
                cff_field(root_cff, field) == cff_field(phase_cff, field)
                for field in ("cff-version", "title", "type", "version")
            )
            and "license:" not in root_cff
            and "license:" not in phase_cff
            and cff_field(root_cff, "version").endswith("-rc.1")
            and "date-released:" not in root_cff
            and "date-released:" not in phase_cff
        ),
        "all_internal_review_reports_and_response_log_exist": all(
            (ROOT / "phase0_5" / "reviews" / name).is_file() for name in review_files
        ),
    }
    archive_report = None
    if args.archive:
        archive_report = verify_archive(args.archive.resolve())
        checks["archive_has_one_safe_regular_tree"] = (
            archive_report["one_top_level_directory"]
            and not archive_report["unsafe_members"]
            and not archive_report["symlink_members"]
        )
        checks["archive_omits_source_study_files"] = not archive_report["forbidden_source_members"]
        checks["archive_checksums_match"] = not archive_report["checksum_failures"]
        checks["archive_source_fingerprint_matches_live_root"] = bool(
            archive_report["source_tree_fingerprint_matches_live_root"]
        )
        checks["archive_package_manifest_matches_live_root"] = not archive_report[
            "live_manifest_failures"
        ]
        clean_report_path = ROOT / "release" / "clean_room_report.json"
        clean_report = (
            json.loads(clean_report_path.read_text()) if clean_report_path.is_file() else {}
        )
        checks["clean_room_report_passes_for_exact_archive"] = (
            clean_report.get("status") == "PASS"
            and clean_report.get("tested_archive", {}).get("sha256") == sha256(args.archive.resolve())
            and clean_report.get("tested_archive", {}).get(
                "source_tree_fingerprint_excluding_runtime_reports"
            )
            == archive_report["package_metadata"].get(
                "source_tree_fingerprint_excluding_generated_outputs"
            )
        )

    checks = {name: bool(value) for name, value in checks.items()}
    failed = [name for name, passed in checks.items() if not passed]
    report = {
        "status": "PASS" if not failed else "FAIL",
        "scope": "technical release readiness only; human gates remain separate",
        "checks": checks,
        "failed": failed,
        "freeze_hash_checks": freeze_checks,
        "scientific_verifier": {
            "status": scientific.get("status"),
            "check_count": scientific.get("check_count"),
        },
        "third_party_cache": {
            "manifest_file_count": len(cache_manifest_rows),
            "alphafold_accessions": cache.get("alphafold_accessions"),
        },
        "workbook": workbook,
        "pdf": {
            "path": str(pdf.relative_to(ROOT)),
            "sha256": sha256(pdf),
            "pages": len(PdfReader(str(pdf)).pages),
        },
        "archive": archive_report,
        "human_gates": {
            "author_signoff": "pending Kyle Nguyen signature",
            "independent_human_methods_review": "pending eligible reviewer",
        },
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, indent=2) + "\n")
    if failed:
        raise SystemExit("technical release verification failed: " + ", ".join(failed))
    print(f"technical release verification PASS: {len(checks)} checks")


if __name__ == "__main__":
    main()
