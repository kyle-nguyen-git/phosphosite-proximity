"""Build release-only audits and the deterministic supplementary workbook.

This script runs after the frozen analysis JSON/CSV artifacts are written. It does not alter the
canonical numerical files named in NUMBERS.md.
"""

from __future__ import annotations

import math
import re
import zipfile
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


HERE = Path(__file__).resolve().parents[1]
ROOT = HERE.parent
RESULTS = HERE / "results"
NUMBERS = ROOT / "NUMBERS.md"
OUTPUT = RESULTS / "robustness_supplement.xlsx"
FIXED_TIMESTAMP = datetime(2026, 7, 29, 0, 0, 0)
HEADER_FILL = PatternFill("solid", fgColor="365F78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="18384D")
PRIMARY_FILL = PatternFill("solid", fgColor="D9EAF5")
INCLUSIVE_FILL = PatternFill("solid", fgColor="FCE6D6")
NOTE_FILL = PatternFill("solid", fgColor="F2F4F5")
LIGHT_BORDER = Border(bottom=Side(style="thin", color="B7CEDD"))


def clean_cell(value: str) -> str:
    return value.replace("**", "").strip()


def parse_arm(text: str, row_prefix: str) -> dict[str, float | int]:
    row = next(line for line in text.splitlines() if line.startswith(f"| {row_prefix}"))
    cells = [clean_cell(cell) for cell in row.strip().strip("|").split("|")]
    interval = re.search(
        r"([0-9.]+)–([0-9.]+) → ([0-9.]+) \[([0-9.]+), ([0-9.]+)\]",
        cells[5],
    )
    if interval is None:
        raise RuntimeError(f"cannot parse canonical arm row from NUMBERS.md: {row}")
    return {
        "n_sites": int(cells[1]),
        "n_proteins": int(cells[2]),
        "n_positive": int(cells[3]),
        "auc": float(cells[4]),
        "ci_low": float(interval.group(1)),
        "ci_high": float(interval.group(2)),
    }


def parse_sift(text: str, row_prefix: str) -> dict[str, float | int]:
    row = next(line for line in text.splitlines() if line.startswith(f"| {row_prefix}"))
    cells = [clean_cell(cell) for cell in row.strip().strip("|").split("|")]
    return {"n_sites": int(cells[1]), "auc": float(cells[3])}


def build_feature_evidence_audit() -> None:
    detailed = pd.read_csv(ROOT / "results" / "uniprot_features_detailed.csv")
    detailed = detailed.loc[detailed.feat_type.isin(["Active site", "Binding site"])].copy()
    detailed["geometry"] = np.where(detailed.start.astype(int) == detailed.end.astype(int), "point", "interval")

    def evidence_class(value: object) -> str:
        if pd.isna(value) or not str(value).strip():
            return "missing"
        codes = set(str(value).split(";"))
        if "ECO:0000269" in codes:
            return "includes_experimental_ECO_0000269"
        if "ECO:0000250" in codes:
            return "includes_sequence_similarity_ECO_0000250"
        return "other_or_mixed"

    detailed["evidence_class"] = detailed.evidence_codes.map(evidence_class)
    detailed["expanded_residue_count"] = detailed.end.astype(int) - detailed.start.astype(int) + 1
    audit = (
        detailed.groupby(["feat_type", "geometry", "evidence_class"], dropna=False, as_index=False)
        .agg(
            feature_records=("feature_id", "size"),
            unique_accessions=("acc", "nunique"),
            expanded_residue_rows=("expanded_residue_count", "sum"),
        )
        .sort_values(["feat_type", "geometry", "evidence_class"])
    )
    audit.to_csv(RESULTS / "feature_evidence_audit.csv", index=False)


def build_replicate_audit() -> None:
    inclusive = pd.read_csv(RESULTS / "robustness_inclusive_sensitivity_analysis.csv")
    included_keys = set(zip(inclusive.acc.astype(str), inclusive.pos.astype(int)))
    members = pd.read_csv(ROOT / "results" / "analysis_site_members.csv")
    phenotype = pd.read_csv(ROOT / "results" / "sites_with_phenotype.csv")
    detail = members[["acc", "pos", "Strain ID"]].merge(
        phenotype[["Strain ID", "raw_n_q05", "has_pheno"]],
        on="Strain ID",
        how="left",
        validate="many_to_one",
    )
    detail = detail.loc[
        [
            (str(acc), int(pos)) in included_keys
            for acc, pos in zip(detail.acc, detail.pos)
        ]
    ].copy()
    rows = []
    for (acc, pos), group in detail.groupby(["acc", "pos"], sort=True):
        group = group.drop_duplicates("Strain ID")
        if len(group) <= 1:
            continue
        positive = group.has_pheno.astype(bool)
        rows.append(
            {
                "acc": acc,
                "pos": int(pos),
                "member_strains": ";".join(group["Strain ID"].astype(str)),
                "n_strains": int(len(group)),
                "n_positive_strains": int(positive.sum()),
                "n_negative_strains": int((~positive).sum()),
                "per_strain_q05_counts": ";".join(
                    str(int(value)) for value in group.raw_n_q05.astype(float)
                ),
                "mean_q05_count": float(group.raw_n_q05.astype(float).mean()),
                "any_positive": bool(positive.any()),
                "unanimous_positive": bool(positive.all()),
                "replicate_discordant": bool(positive.nunique() > 1),
            }
        )
    pd.DataFrame(rows).to_csv(RESULTS / "replicate_aggregation_audit.csv", index=False)


def build_logistic_diagnostics() -> None:
    formulas = (
        "y ~ logd",
        "y ~ logd + plddt + rsa",
        "y ~ logd + plddt + rsa + pae_pair_max",
        "y ~ logd + plddt + rsa + C(pmt_aa_wt)",
        "y ~ logd + plddt + rsa + log_n_annot + log_protein_length + C(pmt_aa_wt)",
    )
    frames = {
        "exclude_annotation_coincident": pd.read_csv(RESULTS / "robustness_primary_analysis.csv"),
        "include_annotation_coincident": pd.read_csv(
            RESULTS / "robustness_inclusive_sensitivity_analysis.csv"
        ),
    }
    rows = []
    for cohort, frame in frames.items():
        frame["y"] = frame.has_pheno.astype(int)
        frame["logd"] = np.log10(frame.dist_core_A + 1.0)
        frame["log_n_annot"] = np.log1p(frame.n_core_residues)
        frame["log_protein_length"] = np.log10(frame.protein_length)
        for formula in formulas:
            fit = smf.logit(formula, data=frame).fit(disp=0)
            design = np.asarray(fit.model.exog, dtype=float)
            outcome = np.asarray(fit.model.endog, dtype=float)
            predicted = np.asarray(fit.predict(), dtype=float)
            weights = np.clip(predicted * (1.0 - predicted), 1e-12, None)
            information_inverse = np.linalg.pinv(design.T @ (weights[:, None] * design))
            leverage = weights * np.einsum(
                "ij,jk,ik->i", design, information_inverse, design
            )
            denominator = np.sqrt(
                np.clip(weights * (1.0 - leverage), 1e-12, None)
            )
            standardized_pearson = (outcome - predicted) / denominator
            rows.append(
                {
                    "cohort": cohort,
                    "formula": formula,
                    "converged": bool(fit.mle_retvals.get("converged", False)),
                    "iterations": int(fit.mle_retvals.get("iterations", -1)),
                    "n_observations": int(fit.nobs),
                    "n_clusters": int(frame.acc.nunique()),
                    "log_likelihood": float(fit.llf),
                    "mcfadden_pseudo_r2": float(fit.prsquared),
                    "min_predicted_probability": float(predicted.min()),
                    "max_predicted_probability": float(predicted.max()),
                    "max_hat_leverage": float(leverage.max()),
                    "max_abs_standardized_pearson_residual": float(
                        np.abs(standardized_pearson).max()
                    ),
                }
            )
    pd.DataFrame(rows).to_csv(RESULTS / "logistic_fit_diagnostics.csv", index=False)


def normalize_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        return None if math.isnan(float(value)) else float(value)
    if pd.isna(value):
        return None
    return value


def number_format(column: str) -> str:
    lowered = column.lower()
    if lowered.startswith("n_") or lowered in {
        "pos", "start", "end", "draws", "iterations", "feature_records",
        "unique_accessions", "expanded_residue_rows", "sequence_start", "sequence_end",
        "uniprot_start", "uniprot_end", "latest_version", "nobs",
    } or "rank" in lowered:
        return "#,##0"
    if "rate" in lowered and "separate" not in lowered:
        return "0.0%"
    if any(
        token in lowered
        for token in (
            "auc", "estimate", "ci_", "rho", "beta", "cluster_p", "pvalue",
            "pseudo_r2", "probability", "leverage", "residual", "distance", "_a",
            "plddt", "rsa", "score", "mean", "median", "brier", "cor_",
            "odds_ratio", "split_",
        )
    ):
        return "0.000"
    return "General"


def safe_table_name(sheet_name: str) -> str:
    return "T_" + re.sub(r"[^A-Za-z0-9_]", "_", sheet_name)


def add_dataframe_sheet(workbook: Workbook, name: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        sheet.append([normalize_value(value) for value in row])

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30

    if len(frame):
        reference = f"A1:{sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate}"
        table = Table(displayName=safe_table_name(name), ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    for column_index, column in enumerate(frame.columns, start=1):
        values = [str(column)] + [
            "" if value is None or (isinstance(value, float) and math.isnan(value)) else str(value)
            for value in frame[column].tolist()
        ]
        width = min(max(max((len(value) for value in values), default=8) + 2, 10), 42)
        if len(frame) > 100:
            width = min(width, 24)
        letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[letter].width = width
        fmt = number_format(str(column))
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=column_index).number_format = fmt
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(
                vertical="top", wrap_text=False
            )


def add_summary(workbook: Workbook, numbers_text: str) -> None:
    primary = parse_arm(numbers_text, "**Primary:")
    inclusive = parse_arm(numbers_text, "**Sensitivity:")
    sift_primary = parse_sift(numbers_text, "**Primary exclude common support**")
    sift_inclusive = parse_sift(numbers_text, "Inclusive common support")
    draw_match = re.search(
        r"Both protein-cluster intervals use \*\*([0-9,]+) draws\*\*",
        numbers_text,
    )
    if draw_match is None:
        raise RuntimeError("cannot parse canonical draw count from NUMBERS.md")
    canonical_draws = int(draw_match.group(1).replace(",", ""))
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "the robustness analysis — two-arm structural-distance calibration"
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells("A2:D2")
    sheet["A2"] = (
        "Primary excludes annotation-coincident substitutions; inclusive retains them at 0 Å"
    )
    sheet["A2"].font = Font(color="6B7E8C", italic=True)
    sheet["A2"].fill = PatternFill("solid", fgColor="E7EFF4")
    rows = [
        ["Metric", "Primary exclude", "Inclusive sensitivity", "Interpretation boundary"],
        ["Substitutions", primary["n_sites"], inclusive["n_sites"], "Primary target population is the exclusion arm."],
        ["Proteins", primary["n_proteins"], inclusive["n_proteins"], "Protein-cluster bootstrap is the reported uncertainty."],
        ["Outcome-positive", primary["n_positive"], inclusive["n_positive"], "All three added inclusive records are outcome-positive."],
        ["Distance AUC", primary["auc"], inclusive["auc"], "Inclusive arm is named in the abstract, not called primary."],
        ["Protein-cluster CI low", primary["ci_low"], inclusive["ci_low"], "One stored interval per arm is reused across artifacts."],
        ["Protein-cluster CI high", primary["ci_high"], inclusive["ci_high"], "The endpoint is a confidence bound, not a utility margin."],
        ["Bootstrap draws", canonical_draws, canonical_draws, "Canonical arm intervals use the same seeded schedule."],
        ["SIFT AUC on common support", sift_primary["auc"], sift_inclusive["auc"], "Post-result comparator; paired inference is reported separately."],
        ["Common-support substitutions", sift_primary["n_sites"], sift_inclusive["n_sites"], "Comparator and distance use common support within each arm."],
    ]
    sheet.row_dimensions[3].height = 8
    for row_index, row in enumerate(rows, start=4):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    for cell in sheet[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index in range(5, 14):
        sheet.cell(row=row_index, column=2).fill = PRIMARY_FILL
        sheet.cell(row=row_index, column=3).fill = INCLUSIVE_FILL
        for column_index in range(1, 5):
            sheet.cell(row=row_index, column=column_index).border = LIGHT_BORDER
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(
                vertical="center", wrap_text=column_index == 4
            )
    for row_index in (5, 6, 7, 11, 13):
        sheet.cell(row=row_index, column=2).number_format = "#,##0"
        sheet.cell(row=row_index, column=3).number_format = "#,##0"
    for row_index in (8, 9, 10, 12):
        sheet.cell(row=row_index, column=2).number_format = "0.000"
        sheet.cell(row=row_index, column=3).number_format = "0.000"
    sheet.merge_cells("A15:D15")
    sheet["A15"] = (
        "Numerical authority: NUMBERS.md. Frozen hashes are in its header; re-verify before changing any numerical claim."
    )
    sheet["A15"].font = Font(color="6B7E8C", italic=True)
    sheet["A15"].fill = NOTE_FILL
    sheet["A15"].alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 64
    table = Table(displayName="T_Summary", ref="A4:D13")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def normalize_xlsx(source: Path, destination: Path) -> None:
    temporary = destination.with_suffix(destination.suffix + ".normalized")
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(
        temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as output:
        for original in sorted(archive.infolist(), key=lambda item: item.filename):
            payload = archive.read(original.filename)
            if original.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2026-07-29T00:00:00Z\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(original.filename, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = original.compress_type
            info.comment = original.comment
            info.extra = b""
            info.create_system = 0
            info.external_attr = original.external_attr
            if original.compress_type == zipfile.ZIP_DEFLATED:
                output.writestr(info, payload, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)
            else:
                output.writestr(info, payload, compress_type=original.compress_type)
    temporary.replace(destination)


def build_workbook() -> None:
    numbers_text = NUMBERS.read_text()
    workbook = Workbook()
    workbook.properties.title = "the robustness analysis structural-distance calibration supplement"
    workbook.properties.subject = "Two-arm exploratory analysis and release audits"
    workbook.properties.creator = "Kyle Nguyen"
    workbook.properties.created = FIXED_TIMESTAMP
    workbook.properties.modified = FIXED_TIMESTAMP
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    add_summary(workbook, numbers_text)

    sheets = [
        ("Arm Estimates", "cohort_arm_primary_estimates.csv"),
        ("Arm Cutoffs", "cohort_arm_cutoffs.csv"),
        ("Arm Descriptives", "cohort_arm_descriptives.csv"),
        ("Confidence Strata", "confidence_strata.csv"),
        ("Regression Models", "regression_models.csv"),
        ("Logistic Diagnostics", "logistic_fit_diagnostics.csv"),
        ("Continuous Outcomes", "continuous_outcomes.csv"),
        ("Predictor Benchmark", "predictor_benchmark.csv"),
        ("Feature Definitions", "feature_definition_sensitivity.csv"),
        ("Feature Evidence", "feature_evidence_audit.csv"),
        ("Primary Analysis Data", "robustness_primary_analysis.csv"),
        ("Inclusive Analysis Data", "robustness_inclusive_sensitivity_analysis.csv"),
        ("Replicate Audit", "replicate_aggregation_audit.csv"),
        ("PAE Manifest", "alphafold_pae_manifest.csv"),
        ("Mismatch Audit", "residue_mismatch_audit.csv"),
        ("Cohort Sensitivity", "cohort_sensitivity.csv"),
        ("PAE at 10A", "pae_column_sensitivity_at_10A.csv"),
        ("PAE Filter Grid", "pae_filter_grid_72x3.csv"),
        ("SIFT Comparator", "sift_comparator_sensitivity.csv"),
        ("Within Protein", "within_protein_auc.csv"),
        ("Residue Classes", "residue_class_sensitivity.csv"),
    ]
    for sheet_name, filename in sheets:
        add_dataframe_sheet(workbook, sheet_name, pd.read_csv(RESULTS / filename))

    temporary = OUTPUT.with_suffix(".xlsx.building")
    workbook.save(temporary)
    normalize_xlsx(temporary, OUTPUT)
    temporary.unlink()
    print(f"wrote {OUTPUT.relative_to(ROOT)} with {len(workbook.sheetnames)} sheets")


def main() -> None:
    build_feature_evidence_audit()
    build_replicate_audit()
    build_logistic_diagnostics()
    build_workbook()


if __name__ == "__main__":
    main()
